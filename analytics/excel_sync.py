from __future__ import annotations

import csv
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

from core.config_manager import BotConfig

log = logging.getLogger(__name__)


class ExcelSync:
    """Actualiza las planillas de interés compuesto y plan de inversión usando los trades del bot."""

    def __init__(
        self,
        config: BotConfig,
        trades_csv: str | Path = "logs/trades/trades.csv",
        interest_path: str | Path = "docs/spreadsheets/12. Interes compuesto.xlsx",
        plan_path: str | Path = "docs/spreadsheets/26. Plan de inversion.xlsx",
    ) -> None:
        self._config = config
        self._trades_csv = Path(trades_csv)
        self._interest_path = Path(interest_path)
        self._plan_path = Path(plan_path)

    def update_all(self) -> None:
        if not self._trades_csv.exists():
            return
        daily_data = self._aggregate_trades()
        if not daily_data:
            return
        try:
            self._update_interest_sheet(daily_data)
        except Exception as exc:  # pragma: no cover - solo logging
            log.warning("No fue posible actualizar '12. Interes compuesto.xlsx': %s", exc)
        try:
            self._update_plan_sheet(daily_data)
        except Exception as exc:  # pragma: no cover
            log.warning("No fue posible actualizar '26. Plan de inversion.xlsx': %s", exc)

    # ------------------------------------------------------------------
    def _aggregate_trades(self) -> List[Dict[str, float]]:
        grouped: Dict[str, Dict[str, float | int]] = {}
        with self._trades_csv.open("r", encoding="utf-8") as fp:
            reader = csv.DictReader(fp)
            for row in reader:
                timestamp = datetime.fromisoformat(row["timestamp"])
                day = timestamp.date().isoformat()
                pnl = float(row["pnl"])
                bucket = grouped.setdefault(
                    day, {"pnl": 0.0, "trades": 0, "wins": 0, "losses": 0}
                )
                bucket["pnl"] += pnl
                bucket["trades"] += 1
                if pnl > 0:
                    bucket["wins"] += 1
                elif pnl < 0:
                    bucket["losses"] += 1

        reference = self._config.risk_limits.reference_account_size_usdt
        ordered_days = sorted(grouped.keys())
        balance = reference
        entries: List[Dict[str, float]] = []
        for day in ordered_days:
            data = grouped[day]
            start_balance = balance
            balance = balance + float(data["pnl"])
            entry = {
                "day": day,
                "capital_start": start_balance,
                "capital_end": balance,
                "pnl": float(data["pnl"]),
                "trades": int(data["trades"]),
                "wins": int(data["wins"]),
                "losses": int(data["losses"]),
            }
            entries.append(entry)
        return entries

    def _update_interest_sheet(self, daily_data: List[Dict[str, float]]) -> None:
        wb = load_workbook(self._interest_path)
        ws = wb.active
        start_row = 5
        risk_pct = self._config.profile.risk_per_trade_pct
        risk_amount = (
            self._config.risk_limits.reference_account_size_usdt * risk_pct
        )
        for idx, data in enumerate(daily_data):
            row = start_row + idx
            ws.cell(row=row, column=3).value = data["day"]
            ws.cell(row=row, column=4).value = round(data["capital_start"], 2)
            ws.cell(row=row, column=5).value = risk_pct
            ws.cell(row=row, column=6).value = round(risk_amount, 2)
            ws.cell(row=row, column=7).value = data["losses"]
            ws.cell(row=row, column=8).value = data["wins"]
            ws.cell(row=row, column=11).value = round(data["pnl"], 2)
        wb.save(self._interest_path)

    def _update_plan_sheet(self, daily_data: List[Dict[str, float]]) -> None:
        wb = load_workbook(self._plan_path)
        ws = wb.active
        target_pct = ws["G3"].value
        if target_pct is None or target_pct == 0:
            target_pct = 0.05

        theoretical_start = self._config.risk_limits.reference_account_size_usdt
        start_row = 5
        for idx, data in enumerate(daily_data):
            row = start_row + idx
            day_date = datetime.fromisoformat(data["day"])
            pnl = data["pnl"]
            start_balance = data["capital_start"]
            end_balance = data["capital_end"]
            theoretical_end = theoretical_start * (1 + target_pct)
            min_goal = theoretical_start * target_pct
            real_goal = start_balance * target_pct
            trades = data["trades"]
            wins = data["wins"]
            winrate = wins / trades if trades else 0

            ws.cell(row=row, column=3).value = day_date
            ws.cell(row=row, column=4).value = 0  # Depositos / retiros
            ws.cell(row=row, column=5).value = round(start_balance, 2)
            ws.cell(row=row, column=6).value = round(theoretical_start, 2)
            ws.cell(row=row, column=7).value = round(start_balance, 2)
            ws.cell(row=row, column=8).value = round(theoretical_end, 2)
            ws.cell(row=row, column=9).value = round(end_balance, 2)
            ws.cell(row=row, column=10).value = round(min_goal, 2)
            ws.cell(row=row, column=11).value = round(real_goal, 2)
            ws.cell(row=row, column=12).value = round(pnl, 2)
            ws.cell(row=row, column=13).value = 1 if pnl >= min_goal else 0
            ws.cell(row=row, column=14).value = trades
            ws.cell(row=row, column=15).value = wins
            ws.cell(row=row, column=16).value = round(winrate, 4)

            theoretical_start = theoretical_end

        wb.save(self._plan_path)

